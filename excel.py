import openpyxl
import os
os.chdir('C:\\Users\\kirst\\OneDrive\\Documents')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
sheet = workbook['Sample 1']
print(type(sheet))
print(workbook.sheetnames)
cell = sheet['A1']
print(cell.value)
print(str(sheet['A1'].value))
print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

